import autoit
import pyinputplus
import openpyxl
import time
import tkinter
from tkinter import filedialog
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys

#prompt user for password, TODO error handling for incorrect password
print("Please enter your password:")
password = pyinputplus.inputPassword() #keeps password input secure

#prompt for digital certificate
root = tkinter.Tk()
root.filename = filedialog.askopenfilename(
    initialdir = "/",
    title = "Select ROS Digital Certificate",
    filetypes = [("", "*.p12.bac")])
certificate = root.filename

#prompt for Client List
root.filename = filedialog.askopenfilename(
    initialdir = "/",
    title = "Select Client List",
    filetypes = [("Excel", "*.xls*")])
client_list = root.filename

#prompt for directory to save TRCs
root.directory = filedialog.askdirectory(
    initialdir = "/",
    title = "Select location to save TRCs",)

#launch browser, go to ROS, upload certificate, and login
browser = webdriver.Firefox()
browser.get('https://www.ros.ie')
time.sleep(5)
linkElem = browser.find_element_by_link_text('Manage My Certificates')
linkElem.click()
time.sleep(2)
browser.find_element_by_xpath(
    '//*[@id="file"]').send_keys(
        certificate.replace("/", '\\'))
time.sleep(3)
browser.find_element_by_xpath(
    '//*[@id="certFilePassword"]').send_keys(
        password)
browser.find_element_by_xpath(
    '//*[@id="submit_importFile"]').send_keys(
        Keys.ENTER)
browser.find_element_by_xpath(
    '/html/body/div[1]/div/div/div[2]/div[1]/div/div/div[2]/form/div[4]/div[1]/div/div[1]/a').send_keys(
        Keys.ENTER)
time.sleep(3)
browser.find_element_by_xpath(
    '//*[@id="password"]').send_keys(
        password)
browser.find_element_by_xpath(
    '//*[@id="password"]').send_keys(
        Keys.ENTER)

#main program loop
wb = openpyxl.load_workbook(client_list) #open client list
sheet = wb['Sheet1']

for x in range (1, sheet.max_row):
    time.sleep(7)
    companyName = sheet.cell(row=x,column=1).value
    browser.find_element_by_xpath(
        '//*[@id="clientName"]').send_keys(
            companyName)
    browser.find_element_by_xpath(
        '//*[@id="clientName"]').send_keys(
            Keys.ENTER)
    time.sleep(7)
    browser.find_element_by_xpath(
        '//*[@id="letterOfResidenceLink"]').send_keys(
            Keys.ENTER)
    time.sleep(7)
    browser.find_element_by_xpath(
        '//*[@id="submitLor"]').send_keys(
            Keys.ENTER)
    time.sleep(7)
    try:
        browser.find_element_by_xpath(
            '//*[@id="confirmLor"]').send_keys(
                Keys.ENTER)
        time.sleep(7)
        browser.find_element_by_xpath(
            '//*[@id="confirmLor"]').send_keys(
                Keys.ENTER)
        time.sleep(7)
        autoit.send('^s')
        time.sleep(7)
        autoit.send(root.directory.replace("/", '\\') + '\\' + companyName)
        time.sleep(7)
        autoit.send('{ENTER}')
        time.sleep(7)
        autoit.send('^w')
        time.sleep(7)
        browser.find_element_by_xpath(
            '//*[@id="greyBtn"]').send_keys(
                Keys.ENTER)
        time.sleep(7)
        browser.find_element_by_xpath(
            '/html/body/div[7]/div[3]/div/button[2]').send_keys(
                Keys.ENTER)
        time.sleep(10)
        browser.find_element_by_xpath(
            '//*[@id="link_id_AGENT_SERVICES"]').send_keys(
                Keys.ENTER)

    except NoSuchElementException:
        print(companyName)
        browser.find_element_by_xpath(
            '//*[@id="greyBtn"]').send_keys(
                Keys.ENTER)
        time.sleep(7)
        autoit.send('{TAB}')
        time.sleep(7)
        autoit.send('{ENTER}')
        time.sleep(10)
        browser.find_element_by_xpath(
            '//*[@id="link_id_AGENT_SERVICES"]').send_keys(
                Keys.ENTER)
